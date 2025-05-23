

import pandas as pd
import streamlit as st
from openpyxl import load_workbook


def append_data_to_excel(data_to_add):
    file_path = 'sales.xlsx'
    sheet_name = 'Sales'
#	Rating	Age																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																													
        
    df1 = pd.DataFrame({
            'Invoice ID': [data_to_add["Invoice ID"]],
            'Branch': [data_to_add["Branch"]],
            'City': [data_to_add["City"]],
            'Customer_type': [data_to_add["Customer_type"]],           
    		"Gender":[data_to_add["Gender"]],
            "Product line":[data_to_add["Product line"]],
            "Unit price":[data_to_add["Unit price"]],
            "Quantity":[data_to_add["Quantity"]],
            "Tax 5%":[data_to_add["Tax 5%"]], 
            "Total":[data_to_add["Total"]],
            "Date":[data_to_add["Date"]],
            "Time":[data_to_add["Time"]],
            "Payment":[data_to_add["Payment"]],
            "cogs":[data_to_add["Cogs"]],
            "gross margin percentage":[data_to_add["Gross margin percentage"]],
            "gross income":[data_to_add["Gross income"]],
            "Rating":[data_to_add["Rating"]],
            "Age":[data_to_add["Age"]], })   

    # load the existing workbook 
    workbook = load_workbook(file_path)
    worksheet = workbook[sheet_name]
    #get the last row in the worksheet
    
    start_row = worksheet.max_row+1  
    #start_row = 5
    # for row in range(5, worksheet.max_row + 2):
    #     if worksheet.cell(row=row, column=2).value is None:
    #         start_row = row
    #         break
    
    workbook.close()
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        
        #writer.sheets = {ws.title: ws for ws in workbook.worksheets}
        df1.to_excel(
            writer,
            sheet_name=sheet_name, 
            index=False,
            header=False,         # don't write the header again
            startrow=start_row-1,# -1 to account for 0-indexing
            startcol=1            # Start writing from the first column B not A
        )
    
    st.success("Data appended successfully!")
    

    
    
    

