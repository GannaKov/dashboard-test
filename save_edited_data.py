from openpyxl import load_workbook
import pandas as pd


def save_edited_data(edited_df, file_path="sales.xlsx", sheet_name="Sales"):
    try:
        # Load existing workbook and remove the sheet if it exists
        workbook = load_workbook(file_path)
        worksheet = workbook[sheet_name]
        # if sheet_name in workbook.sheetnames:
        #     std = workbook[sheet_name]
        #     workbook.remove(std)
        #     workbook.save(file_path)

        

        # Write new sheet with the same name
        with pd.ExcelWriter(file_path, engine='openpyxl', mode='a',if_sheet_exists='overlay') as writer:
            edited_df.to_excel(writer, index=False, startrow=4, 
                               sheet_name=sheet_name, header=False, startcol=1  )
        workbook.close()
        return True, None  # Success

    except Exception as e:
        return False, str(e)  # Error occurred